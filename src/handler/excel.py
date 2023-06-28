import string
from io import BytesIO

from PIL import Image as PImage
from openpyxl.drawing.image import Image
from openpyxl.reader.excel import load_workbook


class Excel:
    def __init__(self, file_path) -> None:
        self.file_path = file_path
        self.work_book = load_workbook(self.file_path)
        self.work_sheet = self.work_book.active

        self.image_row = (self.work_sheet.max_row - 15) // 2
        self.images: dict[str, Image] = {}
        for image in self.work_sheet._images:
            r = image.anchor._from.row + 1
            c = string.ascii_uppercase[image.anchor._from.col]

            self.images[f"{c}{r}"] = image

    @property
    def customer(self):
        return self.work_sheet['D5'].value

    @customer.setter
    def customer(self, text: str):
        self.work_sheet['D5'].value = text

    @property
    def address(self):
        return self.work_sheet['D6'].value

    @address.setter
    def address(self, text: str):
        self.work_sheet['D6'].value = text

    def get_text(self, cell: str):
        col = cell[0]
        row = int(cell[1:])

        return self.work_sheet[f"{col}{row - 1}"].value

    def set_text(self, cell: str, text: str):
        col = cell[0]
        row = int(cell[1:])

        self.work_sheet[f"{col}{row - 1}"].value = text

    def set_image(self, cell: str, image_data: bytes):
        if cell in self.images:
            print(f"{cell} 이미지 삭제")
            self.work_sheet._images.remove(self.images[cell])
            del self.images[cell]

        img = PImage.open(BytesIO(image_data))
        img.resize((1200, 800))
        bytes_io = BytesIO()
        img.save(bytes_io, 'PNG')

        img = Image(bytes_io)
        img.width = 238
        img.height = 160

        self.work_sheet.add_image(img, cell)

    def save(self, file):
        self.work_book.save(file)
        print("Save 완료")
